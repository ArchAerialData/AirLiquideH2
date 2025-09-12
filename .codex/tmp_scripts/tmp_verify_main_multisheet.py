from openpyxl import load_workbook
from pathlib import Path
p = Path(r'test_folders/two_csvs/Combined_Extracted.xlsx')
wb = load_workbook(p)
print('SHEETS', wb.sheetnames)
ws_all = wb['ALL']
print('TAB ALL', getattr(ws_all.sheet_properties.tabColor,'rgb',None))
hdr = [ws_all.cell(row=1, column=c).value for c in range(1,15)]
print('HDR', hdr)
