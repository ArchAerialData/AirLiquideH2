from openpyxl import load_workbook
from pathlib import Path
p = Path(r'test_folders/single_csv/Combined_Extracted.xlsx')
wb = load_workbook(p)
ws = wb.active
# header row
hdr = [ws.cell(row=2, column=c).value for c in range(1,15)]
print('HEADER', hdr)
# sample 10 rows: show PLID, BeginMeasu, EndMeasure, Route_Desc, Class_Loca, Diameter, Product
for r in range(3, 13):
    vals = [ws.cell(row=r, column=c).value for c in (2,4,5,6,7,8,9)]
    print('KMZ ROW', r, vals)
