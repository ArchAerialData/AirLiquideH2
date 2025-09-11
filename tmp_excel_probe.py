from openpyxl import load_workbook
from pathlib import Path
p = Path('.codex/test-extract/raw-csv/Combined_Extracted.xlsx')
wb = load_workbook(p)
ws = wb.active
# dump first 12 rows
for r in range(1, 13):
    row = [ws.cell(row=r, column=c).value for c in range(1, 20)]
    if any(v is not None for v in row):
        print(r, row)
# try to detect a header row (row with multiple non-None values not the merged project row)
hdr_row_idx = None
for r in range(1, 12):
    vals = [ws.cell(row=r, column=c).value for c in range(1, 20)]
    if sum(1 for v in vals if v) >= 4:
        hdr_row_idx = r
        break
print('hdr_row_idx', hdr_row_idx)
if hdr_row_idx:
    fills = []
    for c in range(1, 20):
        cell = ws.cell(row=hdr_row_idx, column=c)
        fill = getattr(cell.fill, 'fgColor', None)
        fills.append(getattr(fill, 'rgb', None))
    print('fills', fills)
