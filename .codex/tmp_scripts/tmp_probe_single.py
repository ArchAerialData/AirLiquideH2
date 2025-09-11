from openpyxl import load_workbook
from pathlib import Path
p = Path(r'test_folders/single_csv/Combined_Extracted.xlsx')
wb = load_workbook(p)
ws = wb.active
# find header row (first with many non-empty)
hdr_row = None
for r in range(1, 10):
    vals = [ws.cell(row=r, column=c).value for c in range(1, 15)]
    if sum(1 for v in vals if v) >= 5:
        hdr_row = r
        print('HEADER', r, vals[:14])
        break
# print first 5 data rows under header
if hdr_row:
    for r in range(hdr_row+1, hdr_row+6):
        vals = [ws.cell(row=r, column=c).value for c in range(1, 15)]
        print('ROW', r, vals[:14])
