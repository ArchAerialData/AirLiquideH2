from openpyxl import load_workbook
from pathlib import Path
p = Path(r'test_folders/single_csv/Combined_Extracted.xlsx')
wb = load_workbook(p)
print('SHEETS', wb.sheetnames)
ws_all = wb['ALL']
print('TAB ALL', getattr(ws_all.sheet_properties.tabColor,'rgb',None))
hdr = [ws_all.cell(row=1, column=c).value for c in range(1,20)]
print('HDR', hdr)
# count rows
data_rows = 0
for r in range(2, ws_all.max_row+1):
    if any(ws_all.cell(row=r, column=c).value is not None for c in range(1, len(hdr)+1)):
        data_rows += 1
print('ALL_ROWS', data_rows)
