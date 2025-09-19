from __future__ import annotations

from pathlib import Path
import shutil
from typing import Iterable, List, Tuple

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

from .config import SeparatorStyle
from .output_schema import OUTPUT_COLUMNS, CSV_HEADERS, KMZ_HEADERS, TRAILING_EXTRAS


BLUE = "FF366092"   # header for CSV fields
GOLD = "FFC9AD6A"   # header for KMZ fields
ALL_TAB_YELLOW = "FFFFFF00"  # sheet tab color for ALL


def _month_label(m: int) -> str:
    mapping = {1: "Jan", 2: "Feb", 3: "Mar", 4: "Apr", 5: "May", 6: "Jun", 7: "Jul", 8: "Aug", 9: "Sept", 10: "Oct", 11: "Nov", 12: "Dec"}
    return mapping.get(m, str(m))


def _date_to_sheet_name(d) -> str:
    # d can be a datetime.date or datetime
    try:
        y = d.year
        m = d.month
        day = d.day
        name = f"{_month_label(m)} {day}, {y}"
    except Exception:
        name = "Unknown Date"
    # sanitize for Excel
    for ch in ":\\/?*[]":
        name = name.replace(ch, "-")
    return name[:31]


def _ensure_unique_sheet_name(wb: Workbook, base: str) -> str:
    name = base
    i = 2
    while name in wb.sheetnames:
        suffix = f" ({i})"
        name = (base + suffix)[:31]
        i += 1
    return name


def _style_header(ws, headers: List[str]) -> None:
    header_font = Font(color="FFFFFF", bold=True)
    ws.append(headers)
    row_idx = 1
    for idx, col_name in enumerate(headers, start=1):
        cell = ws.cell(row=row_idx, column=idx)
        if col_name in CSV_HEADERS:
            cell.fill = PatternFill(fill_type="solid", fgColor=BLUE)
        elif col_name in KMZ_HEADERS:
            cell.fill = PatternFill(fill_type="solid", fgColor=GOLD)
        else:
            # Default any additional columns (e.g., diagnostics) to BLUE header styling
            cell.fill = PatternFill(fill_type="solid", fgColor=BLUE)
        cell.font = header_font
        # Header alignment: center by default; Route_Desc may be left aligned if desired
        if col_name == "Route_Desc":
            cell.alignment = Alignment(horizontal="left")
        else:
            cell.alignment = Alignment(horizontal="center")


def _sanitize_filename(text: str) -> str:
    return "".join(ch if ch.isalnum() or ch in ("-", "_", ".") else "_" for ch in str(text).strip())


def _kml_for_point(name: str, lat: float, lon: float, description_html: str | None = None) -> str:
    desc = description_html or ""
    return f"""<?xml version=\"1.0\" encoding=\"UTF-8\"?>
<kml xmlns=\"http://www.opengis.net/kml/2.2\">
  <Document>
    <Placemark>
      <name>{name}</name>
      <description><![CDATA[{desc}]]></description>
      <Point>
        <coordinates>{lon},{lat},0</coordinates>
      </Point>
    </Placemark>
  </Document>
</kml>
"""


def _write_data(ws, df: pd.DataFrame, headers: List[str], kml_dir: Path) -> Tuple[int, int]:
    # Start writing right after header (row 2)
    start_row = 2
    view_idx = headers.index("View in Google Earth") + 1 if "View in Google Earth" in headers else None
    lat_name = "Latitude" if "Latitude" in headers else None
    lon_name = "Longitude" if "Longitude" in headers else None

    for i in range(len(df)):
        def _to_excel(v):
            try:
                if pd.isna(v):
                    return None
            except Exception:
                pass
            return v
        row_vals = [_to_excel(df.iloc[i].get(col)) for col in headers]
        ws.append(row_vals)
        excel_row = ws.max_row
        # Hyperlink generation per row
        if view_idx and lat_name and lon_name:
            try:
                lat = float(df.iloc[i][lat_name])
                lon = float(df.iloc[i][lon_name])
                if pd.notna(lat) and pd.notna(lon):
                    kml_dir.mkdir(parents=True, exist_ok=True)
                    # Name KMLs using the PPM value (rounded to 2 decimals) and add a unique row suffix
                    ppm_val = df.iloc[i].get("PPM")
                    if pd.isna(ppm_val):
                        base = f"row_{i+1}"
                        ppm_disp = ""
                    else:
                        try:
                            ppm_float = float(ppm_val)
                            ppm_disp = f"{ppm_float:.2f}"
                        except Exception:
                            ppm_disp = str(ppm_val)
                        base = f"PPM_{ppm_disp}_r{i+1}"
                    base = _sanitize_filename(base)
                    kml_path = kml_dir / f"{base}.kml"
                    # Build HTML with client-like styling (header band + inner table with zebra rows)
                    def val(name: str):
                        return "" if name not in headers else ("" if pd.isna(df.iloc[i].get(name)) else str(df.iloc[i].get(name)))
                    plid = val("PLID")
                    bme = val("BeginMeasu")
                    eme = val("EndMeasure")
                    header_title = f"{plid} {bme} - {eme}".strip()
                    # Format coords to 7 decimals for display
                    try:
                        lat_disp = f"{float(df.iloc[i].get('Latitude')):.7f}"
                    except Exception:
                        lat_disp = val("Latitude")
                    try:
                        lon_disp = f"{float(df.iloc[i].get('Longitude')):.7f}"
                    except Exception:
                        lon_disp = val("Longitude")
                    # PPM heading centered with suffix (rounded to 2 decimals)
                    ppm_text = "" if pd.isna(ppm_val) else f"{ppm_disp} PPM"
                    rows_kv = [
                        ("PPM", ppm_text.replace(" PPM", "")),
                        ("PLID", plid),
                        ("BeginMeasu", bme),
                        ("EndMeasure", eme),
                        ("Route_Desc", val("Route_Desc")),
                        ("Class_Loca", val("Class_Loca")),
                        ("Diameter", val("Diameter")),
                        ("Product", val("Product")),
                        ("Date", val("Date")),
                        ("Longitude", lon_disp),
                        ("Latitude", lat_disp),
                        ("Serial No.", val("Serial No.")),
                        ("Source_File", val("Source_File")),
                    ]
                    outer = []
                    outer.append('<html xmlns:fo="http://www.w3.org/1999/XSL/Format" xmlns:msxsl="urn:schemas-microsoft-com:xslt">')
                    outer.append('<head><META http-equiv="Content-Type" content="text/html"></head>')
                    outer.append('<body style="margin:0;overflow:auto;background:#FFFFFF;">')
                    outer.append('<table style="font-family:Arial,Verdana,Times;font-size:12px;text-align:left;width:100%;border-collapse:collapse;padding:3px 3px 3px 3px">')
                    # (Dedup) Do not add a centered PPM line in the body; we'll use the KML <name> instead
                    # Header band
                    outer.append('<tr style="text-align:center;font-weight:bold;background:#9CBCE2"><td>')
                    outer.append(header_title)
                    outer.append('</td></tr>')
                    # Inner table
                    outer.append('<tr><td><table style="font-family:Arial,Verdana,Times;font-size:12px;text-align:left;width:100%;border-spacing:0px; padding:3px 3px 3px 3px">')
                    for idx_row, (label, value) in enumerate(rows_kv):
                        zebra = " bgcolor=\"#D4E4F3\"" if (idx_row % 2 == 1) else ""
                        outer.append(f"<tr{zebra}><td style='font-weight:bold;padding:2px 10px 2px 8px'>{label}</td><td style='padding:2px 8px'>{value}</td></tr>")
                    outer.append('</table></td></tr></table></body></html>')
                    # Only write the file if it does not already exist; otherwise reuse existing file
                    if not kml_path.exists():
                        # Use the KML <name> to show e.g. "3.51 PPM" in the balloon header and above the pin
                        kml_name = ppm_text if ppm_text else base
                        kml_path.write_text(_kml_for_point(kml_name, lat, lon, "".join(outer)), encoding="utf-8")
                    cell = ws.cell(row=excel_row, column=view_idx)
                    cell.value = "View Placemark"
                    cell.hyperlink = str(kml_path.resolve())
                    cell.font = Font(color="0000EE", underline="single")
            except Exception:
                pass

    end_row = ws.max_row
    return start_row, end_row


def _apply_number_formats(ws, headers: List[str], start_row: int, end_row: int) -> None:
    # Date
    if "Date" in headers:
        c_idx = headers.index("Date") + 1
        for r in range(start_row, end_row + 1):
            cell = ws.cell(row=r, column=c_idx)
            if cell.value:
                cell.number_format = "mm/dd/yyyy"

    # PLID integer
    if "PLID" in headers:
        c_idx = headers.index("PLID") + 1
        for r in range(start_row, end_row + 1):
            cell = ws.cell(row=r, column=c_idx)
            if isinstance(cell.value, (int, float)):
                cell.number_format = "0"

    # Begin/End two decimals
    for name in ("BeginMeasu", "EndMeasure"):
        if name in headers:
            c_idx = headers.index(name) + 1
            for r in range(start_row, end_row + 1):
                cell = ws.cell(row=r, column=c_idx)
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "0.00"

    # Distance diagnostics
    if "Nearest_Distance_Meters" in headers:
        c_idx = headers.index("Nearest_Distance_Meters") + 1
        for r in range(start_row, end_row + 1):
            cell = ws.cell(row=r, column=c_idx)
            if isinstance(cell.value, (int, float)):
                cell.number_format = "0.00"


def _apply_alignments(ws, headers: List[str], start_row: int, end_row: int) -> None:
    # Gold columns centered, except Route_Desc left
    targets: List[Tuple[int, str]] = []
    for col in KMZ_HEADERS:
        if col in headers:
            align = "left" if col == "Route_Desc" else "center"
            targets.append((headers.index(col) + 1, align))
    # Center the hyperlink column as well
    if "View in Google Earth" in headers:
        targets.append((headers.index("View in Google Earth") + 1, "center"))
    for r in range(start_row, end_row + 1):
        for c_idx, align in targets:
            cell = ws.cell(row=r, column=c_idx)
            if align == "left":
                cell.alignment = Alignment(horizontal="left")
            else:
                cell.alignment = Alignment(horizontal="center")


def _apply_ppm_conditional(ws, headers: List[str], start_row: int, end_row: int) -> None:
    if "PPM" not in headers or end_row < start_row:
        return
    c_idx = headers.index("PPM") + 1
    col_letter = get_column_letter(c_idx)
    cell_range = f"{col_letter}{start_row}:{col_letter}{end_row}"
    fill = PatternFill(start_color=ALL_TAB_YELLOW, end_color=ALL_TAB_YELLOW, fill_type="solid")
    rule = CellIsRule(operator='greaterThanOrEqual', formula=['5'], fill=fill)
    ws.conditional_formatting.add(cell_range, rule)


def _apply_distance_row_conditional(ws, headers: List[str], start_row: int, end_row: int) -> None:
    # Shade entire rows light grey when Nearest_Distance_Meters > 500
    if "Nearest_Distance_Meters" not in headers or end_row < start_row:
        return
    dist_idx = headers.index("Nearest_Distance_Meters") + 1
    dist_col = get_column_letter(dist_idx)
    last_col = get_column_letter(len(headers))
    apply_range = f"A{start_row}:{last_col}{end_row}"
    formula = f"INDIRECT(\"{dist_col}\"&ROW())>500"
    fill = PatternFill(start_color="FFEFEFEF", end_color="FFEFEFEF", fill_type="solid")
    from openpyxl.formatting.rule import FormulaRule
    rule = FormulaRule(formula=[formula], fill=fill)
    ws.conditional_formatting.add(apply_range, rule)


def _autosize_columns(ws, headers: List[str]) -> None:
    for col_idx in range(1, len(headers) + 1):
        letter = get_column_letter(col_idx)
        max_len = 0
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
            v = row[0].value
            if v is not None:
                max_len = max(max_len, len(str(v)))
        ws.column_dimensions[letter].width = min(max_len + 2, 50)


def _add_sheet_with_df(wb: Workbook, name: str, df: pd.DataFrame, headers: List[str], tab_color: str | None = None):
    ws = wb.create_sheet(title=name)
    if tab_color:
        ws.sheet_properties.tabColor = tab_color
    _style_header(ws, headers)
    kml_dir = Path("temp_kml")
    start_row, end_row = _write_data(ws, df, headers, kml_dir)
    _apply_number_formats(ws, headers, start_row, end_row)
    _apply_alignments(ws, headers, start_row, end_row)
    _apply_ppm_conditional(ws, headers, start_row, end_row)
    _apply_distance_row_conditional(ws, headers, start_row, end_row)
    _autosize_columns(ws, headers)
    return ws


def write_all_and_date_sheets(
    combined_df: pd.DataFrame,
    output_path: Path,
    separator_style: SeparatorStyle | None = None,
) -> None:
    """Write a workbook with an ALL sheet and one sheet per unique Date value.

    combined_df is assumed to already be in final output schema (OUTPUT_COLUMNS order).
    """
    # Determine headers: base + any known trailing extras present
    df = combined_df.copy()
    extras = [c for c in TRAILING_EXTRAS if c in df.columns]
    headers = OUTPUT_COLUMNS + extras

    # Clean temp KML directory before each run
    tmp_dir = Path("temp_kml")
    try:
        if tmp_dir.exists():
            shutil.rmtree(tmp_dir)
    except Exception:
        pass

    wb = Workbook()
    # Remove default sheet
    default = wb.active
    wb.remove(default)

    # ALL sheet first
    _add_sheet_with_df(wb, _ensure_unique_sheet_name(wb, "ALL"), df, headers, tab_color=ALL_TAB_YELLOW)

    # Per-date sheets
    if "Date" in df.columns:
        # Normalize Date column to datetime.date for grouping
        dates = pd.to_datetime(df["Date"], errors="coerce").dt.date
        df = df.copy()
        df["__DATE_ONLY__"] = dates
        unique_dates = sorted([d for d in dates.dropna().unique().tolist()])
        for d in unique_dates:
            name = _date_to_sheet_name(d)
            name = _ensure_unique_sheet_name(wb, name)
            dmask = df["__DATE_ONLY__"] == d
            df_date = df.loc[dmask, headers].copy()
            _add_sheet_with_df(wb, name, df_date, headers)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
