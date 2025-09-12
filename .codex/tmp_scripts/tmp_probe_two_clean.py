from openpyxl import load_workbook
from pathlib import Path
p = Path(r'test_folders/two_csvs/Combined_Extracted_Clean.xlsx')
wb = load_workbook(p)
ws = wb.active
# Find merged header rows and per-group counts
rows = ws.max_row
cols = ws.max_column
# Capture lines that start with project name banner
banners = []
for r in range(1, rows+1):
    v = ws.cell(row=r, column=1).value
    if isinstance(v, str) and v.strip().startswith('--- ProjectName'):
        banners.append(r)
print('BANNERS', banners)
# Print headers and first few data rows after each banner
for i, start in enumerate(banners):
    hdr_row = start + 1
    hdr = [ws.cell(row=hdr_row, column=c).value for c in range(1, 15)]
    print(f'GROUP {i+1} HEADER', hdr)
    # data until next banner or end
    data_start = hdr_row + 1
    data_end = (banners[i+1]-1) if i+1 < len(banners) else rows
    count = 0
    for r in range(data_start, min(data_start+5, data_end+1)):
        vals = [ws.cell(row=r, column=c).value for c in range(1, 15)]
        print('ROW', r, vals)
        count += 1
    print('GROUP_ROWS_RANGE', data_start, data_end, 'SAMPLE_COUNT', count)
