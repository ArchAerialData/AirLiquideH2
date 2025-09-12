from openpyxl import load_workbook
from pathlib import Path
p = Path(r'test_folders/two_csvs/Combined_Extracted_Clean.xlsx')
wb = load_workbook(p)
print('SHEETS', wb.sheetnames)
ws_all = wb['ALL']
print('TAB ALL', getattr(ws_all.sheet_properties.tabColor,'rgb',None))
# Check header row in ALL
hdr = [ws_all.cell(row=1, column=c).value for c in range(1,15)]
print('HDR', hdr)
# Find PPM column letter
ppm_idx = hdr.index('PPM')+1
print('PPM_COL', ppm_idx)
