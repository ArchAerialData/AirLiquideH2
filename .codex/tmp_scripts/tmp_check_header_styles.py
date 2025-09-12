from openpyxl import load_workbook
from pathlib import Path
p = Path(r'test_folders/two_csvs/Combined_Extracted_Clean.xlsx')
wb = load_workbook(p)
print('SHEETS', wb.sheetnames)
ws = wb['ALL']
hdr = [ws.cell(row=1, column=c).value for c in range(1, 20)]
print('HDR', hdr)
# Read header fill RGB for the last two columns
from openpyxl.styles import PatternFill
for name in ('KMZ_Match_Method', 'Nearest_Distance_Meters'):
    try:
        idx = hdr.index(name)+1
    except ValueError:
        print('MISSING', name)
        continue
    cell = ws.cell(row=1, column=idx)
    fill = getattr(cell.fill,'fgColor',None)
    print('FILL', name, getattr(fill,'rgb',None))
