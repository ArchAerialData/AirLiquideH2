from openpyxl import load_workbook
from pathlib import Path
p = Path(r'test_folders/two_csvs/Combined_Extracted_Clean.xlsx')
wb = load_workbook(p)
print('SHEETS', wb.sheetnames)
for name in wb.sheetnames:
    ws = wb[name]
    col = getattr(ws.sheet_properties, 'tabColor', None)
    rgb = getattr(col, 'rgb', None) if col else None
    print('TAB', name, rgb)
# Inspect the first sheet header
ws0 = wb[wb.sheetnames[0]]
h1 = [ws0.cell(row=1, column=c).value for c in range(1, 20)]
print('HDR_ROW1', h1)
# Inspect another sheet if exists
if len(wb.sheetnames) > 1:
    ws1 = wb[wb.sheetnames[1]]
    h2 = [ws1.cell(row=1, column=c).value for c in range(1, 20)]
    print('HDR_ROW1_SHEET2', h2)
