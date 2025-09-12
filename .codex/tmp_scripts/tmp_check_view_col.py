from openpyxl import load_workbook
from pathlib import Path
p = Path(r'test_folders/two_csvs/Combined_Extracted_Clean.xlsx')
wb = load_workbook(p)
ws = wb['ALL']
hdr = [ws.cell(row=1, column=c).value for c in range(1, 20)]
print('HDR', hdr)
# check the hyperlink exists in the View column
try:
    view_idx = hdr.index('View in Google Earth')+1
    print('VIEW_COL', view_idx)
    for r in range(2, ws.max_row+1):
        cell = ws.cell(row=r, column=view_idx)
        if cell.value:
            print('ROW', r, 'TEXT', cell.value, 'HYPER', cell.hyperlink.target if cell.hyperlink else None)
            break
except ValueError:
    print('NO_VIEW_COL')
