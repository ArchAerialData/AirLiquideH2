from __future__ import annotations

from pathlib import Path
from typing import List

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font, Alignment

from .aggregator import GroupBoundary
from .config import SeparatorStyle
from .output_schema import OUTPUT_COLUMNS, CSV_HEADERS, KMZ_HEADERS


def write_with_separators(
    combined_df: pd.DataFrame,
    boundaries: List[GroupBoundary],
    output_path: Path,
    separator_style: SeparatorStyle,
) -> None:
    """Write XLSX matching the desired grouped layout:
    - Grey merged row per group with project/csv and row count
    - Repeated header row per group with dual colors (blue for CSV, gold for KMZ)
    - Data rows following
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Combined Flight Data"

    # Colors (tweak if you prefer different hex)
    BLUE = "FF366092"   # header for CSV fields
    GOLD = "FFC9AD6A"   # header for KMZ fields
    GREY = "FF4D4D4D"   # merged project row fill

    header_font = Font(color="FFFFFF", bold=True)
    italic_white = Font(color="FFFFFFFF", italic=True)

    # Ensure combined_df only has the expected columns and order
    headers = OUTPUT_COLUMNS

    current_row = 1
    for boundary in boundaries:
        # 1) Grey merged row across the full header width
        text = f"--- ProjectName : {boundary.project_name} Organization :  ({boundary.csv_name}) - {boundary.row_count} rows ---"
        ws.cell(row=current_row, column=1).value = text
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(headers))
        mc = ws.cell(row=current_row, column=1)
        mc.fill = PatternFill(fill_type="solid", fgColor=GREY)
        mc.font = italic_white
        mc.alignment = Alignment(horizontal="center")
        current_row += 1

        # 2) Header row with blue/gold sections
        ws.append(headers)
        for idx, col_name in enumerate(headers, start=1):
            cell = ws.cell(row=current_row, column=idx)
            if col_name in CSV_HEADERS:
                cell.fill = PatternFill(fill_type="solid", fgColor=BLUE)
            elif col_name in KMZ_HEADERS:
                cell.fill = PatternFill(fill_type="solid", fgColor=GOLD)
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        current_row += 1

        # 3) Write group rows in order
        csv_data = combined_df.iloc[boundary.start_row : boundary.end_row + 1][headers]
        start_data_row = current_row
        for row_data in dataframe_to_rows(csv_data, index=False, header=False):
            ws.append(row_data)
            current_row += 1

        # 4) Format Date column as mm/dd/yyyy for this group's rows
        try:
            date_col_idx = headers.index("Date") + 1
            for r in range(start_data_row, current_row):
                cell = ws.cell(row=r, column=date_col_idx)
                # Only set for date-like values
                if cell.value:
                    cell.number_format = "mm/dd/yyyy"
        except Exception:
            pass

    # Auto-size columns
    from openpyxl.utils import get_column_letter
    for col_idx in range(1, len(headers) + 1):
        max_length = 0
        letter = get_column_letter(col_idx)
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
            cell = row[0]
            try:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[letter].width = min(max_length + 2, 50)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
