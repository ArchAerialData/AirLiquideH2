from openpyxl import load_workbook
from pathlib import Path
p = Path(r'test_folders/two_csvs/Combined_Extracted_Clean.xlsx')
wb = load_workbook(p)
print('SHEETS', wb.sheetnames)
ws = wb['ALL']
hdr = [ws.cell(row=1, column=c).value for c in range(1, 20)]
print('HDR', hdr)
# find extras positions
try:
    d_idx = hdr.index('Nearest_Distance_Meters')+1
except ValueError:
    d_idx = None
print('DIST_COL', d_idx)
# print first 10 data rows
for r in range(2, min(ws.max_row, 12)):
    vals = [ws.cell(row=r, column=c).value for c in range(1, len(hdr)+1)]
    print('ROW', r, vals)
