from openpyxl import load_workbook
from pathlib import Path
p = Path(r'test_folders/two_csvs/Combined_Extracted_Clean.xlsx')
wb = load_workbook(p)
ws = wb['ALL']
hdr = [ws.cell(row=1, column=c).value for c in range(1, 20)]
view_idx = hdr.index('View in Google Earth')+1
# Check alignment of a data cell in view column
print('VIEW_COL', view_idx, 'ALIGN', ws.cell(row=2, column=view_idx).alignment.horizontal)
print('TEXT', ws.cell(row=2, column=view_idx).value)
print('HYPER', ws.cell(row=2, column=view_idx).hyperlink.target if ws.cell(row=2, column=view_idx).hyperlink else None)
# Inspect a KML file content
import os
from glob import glob
files = sorted(glob('temp_kml/*.kml'))
print('KML_COUNT', len(files))
if files:
    fp = files[0]
    print('SAMPLE_KML', fp)
    print(Path(fp).read_text(encoding='utf-8')[:200].replace('\n',' '))
