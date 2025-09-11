from openpyxl import load_workbook
from pathlib import Path
p = Path(r'test_folders/single_csv/Combined_Extracted.xlsx')
wb = load_workbook(p)
ws = wb.active
# verify alignments for a few cells
headers = [ws.cell(row=2, column=c).value for c in range(1,15)]
print('HDRS', headers)
# Get alignment of header Route_Desc and Class_Loca
import sys
rd_idx = headers.index('Route_Desc')+1
cl_idx = headers.index('Class_Loca')+1
print('Header Route_Desc align:', ws.cell(row=2, column=rd_idx).alignment.horizontal)
print('Header Class_Loca  align:', ws.cell(row=2, column=cl_idx).alignment.horizontal)
# Check first data row alignments
r=3
def a(cname):
    return ws.cell(row=r, column=headers.index(cname)+1).alignment.horizontal
print('Row3 align PLID, Begin, End, Class_Loca, Diameter, Product, Route_Desc:')
print(a('PLID'), a('BeginMeasu'), a('EndMeasure'), a('Class_Loca'), a('Diameter'), a('Product'), a('Route_Desc'))
