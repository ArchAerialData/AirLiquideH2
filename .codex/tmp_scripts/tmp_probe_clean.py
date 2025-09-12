from openpyxl import load_workbook
from pathlib import Path
p = Path(r'test_folders/single_csv/Combined_Extracted_Clean.xlsx')
wb = load_workbook(p)
ws = wb.active
# header row
hdr = [ws.cell(row=2, column=c).value for c in range(1,15)]
print('HEADER', hdr)
rows=0
for r in range(3, 40):
    vals = [ws.cell(row=r, column=c).value for c in range(1,15)]
    if any(v is not None for v in vals):
        rows+=1
        print('ROW', r, vals)
print('DATA_ROWS', rows)
