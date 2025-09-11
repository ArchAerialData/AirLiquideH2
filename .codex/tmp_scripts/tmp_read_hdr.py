from openpyxl import load_workbook
from pathlib import Path
p = Path(r'test_folders/single_csv/Combined_Extracted.xlsx')
wb = load_workbook(p)
ws = wb.active
hdr = [ws.cell(row=2, column=c).value for c in range(1, 50)]
hdr = [h for h in hdr if h is not None]
print(hdr)
